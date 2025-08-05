import openpyxl
import json
import os

from constants import PAGE_METADATA,TABS_METADATA



def excel_to_json():
    try:
        # Get the directory of the current script
        script_dir = os.path.dirname(os.path.abspath(__file__))
        print(script_dir)
        # Define the path to the Excel file
        file_path = os.path.join(script_dir, "doc.xlsx")
        # Define the path to the JSON file
        json_path = os.path.join(script_dir, "public/assets", "landing-page.json")


        # Open the Excel file
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        try:
            sheet = workbook[PAGE_METADATA["HOME_PAGE"]]
        except KeyError:
            print("Error: Sheet 'Data on homepage' not found in the Excel file.")
            print(f"Available sheets: {workbook.sheetnames}")
            return


        # Get headers
        headers = [cell.value for cell in sheet[1]]
        cleaned_headers = [str(cell).strip() if cell is not None else '' for cell in headers]
        expected_columns = TABS_METADATA["HOME_PAGE"]
        if not all(col in cleaned_headers for col in expected_columns):
            print(f"Error: Excel file must contain columns: {expected_columns}")
            print(f"Found: {cleaned_headers}")
            return

        # Extract data rows
        data = []
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            try:
                row_data = {
                    'label': row[cleaned_headers.index(TABS_METADATA["HOME_PAGE"][0])] or '',
                    'value': row[cleaned_headers.index(TABS_METADATA["HOME_PAGE"][2])] or '',
                    'icon': row[cleaned_headers.index(TABS_METADATA["HOME_PAGE"][3])] or ''
                }
                if isinstance(row_data['value'], float) and row_data['value'].is_integer():
                    row_data['value'] = int(row_data['value'])
                data.append(row_data)
            except Exception as e:
                print(f"Error processing row {row_idx}: {str(e)}")
                continue


        # Read the existing JSON file
        with open(json_path, 'r', encoding='utf-8') as json_file:
            raw_content = json_file.read()
            json_file.seek(0)
            try:
                json_data = json.load(json_file)
            except json.JSONDecodeError:
                try:
                    json_data = json.loads(raw_content)
                except json.JSONDecodeError:
                    json_data = []

        # Ensure json_data is a list
        if not isinstance(json_data, list):
            json_data = [json_data] if json_data else []

        # Find and update the "data-indicators" object
        found = False
        for obj in json_data:
            if isinstance(obj, dict) and obj.get('type', '').strip().lower() == 'data-indicators':
                obj['indicators'] = data
                found = True
                break

        # If not found, append (this is a fallback, but your JSON should already have it)
        if not found:
            json_data.append({'type': 'data-indicators', 'indicators': data})


        # Write back to the file
        with open(json_path, 'w', encoding='utf-8') as json_file:
            json.dump(json_data, json_file, indent=2, ensure_ascii=False)


    except Exception as e:
        print(f"Error: {str(e)}")

if __name__ == "__main__":
    excel_to_json()
