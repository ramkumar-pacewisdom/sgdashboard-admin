import openpyxl
import json
import os
import importlib.util
from constants import PAGE_METADATA, TABS_METADATA

def load_state_codes():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    state_codes_file = os.path.join(script_dir, "..", "pages", "state_code_details.json")
    if not os.path.exists(state_codes_file):
        print("❌ state_code_details.json not found.")
        return None
    with open(state_codes_file, "r", encoding="utf-8") as f:
        return json.load(f)

def format_metric_value(val):
    """Ensure clean string values without unnecessary .0"""
    if isinstance(val, (int, float)):
        if float(val).is_integer():
            return str(int(val))
        else:
            return f"{val:.2f}".rstrip("0").rstrip(".")
    return str(val).strip()


def extract_district_details(excel_file):
    try:
        state_codes = load_state_codes()
        if not state_codes:
            return

        workbook = openpyxl.load_workbook(excel_file, data_only=True)

        try:
            sheet = workbook[PAGE_METADATA["DISTRICT_DETAILS"]]
        except KeyError:
            print(f"❌ Sheet not found: {PAGE_METADATA['DISTRICT_DETAILS']}")
            return

        expected_headers = TABS_METADATA["DISTRICT_DETAILS"]
        column_indices = {}
        for cell in sheet[1]:
            if cell.value and str(cell.value).strip() in expected_headers:
                column_indices[str(cell.value).strip()] = cell.column

        missing_columns = [col for col in expected_headers if col not in column_indices]
        if missing_columns:
            print(f"❌ Missing required columns: {missing_columns}")
            return

        # Exclusions
        excluded_codes = ["categories", "active missions", "community led initiative", "state led program", "district led program" ]
        excluded_for_metrics = ["categories", "community led initiative", "state led program", "district led program"]

        # State-level container
        states_map = {}

        # ✅ District-level containers
        district_files_map = {}

        row_num = 2
        while True:
            state_name = sheet.cell(row=row_num, column=column_indices["State Name"]).value
            if not state_name:
                break

            district_name = sheet.cell(row=row_num, column=column_indices["District Name"]).value
            indicator = sheet.cell(row=row_num, column=column_indices["Indicator"]).value or ""
            definition = sheet.cell(row=row_num, column=column_indices["Definition"]).value or ""
            data_value = sheet.cell(row=row_num, column=column_indices["Data"]).value

            state_name = str(state_name).strip()
            district_name = str(district_name).strip()
            indicator = str(indicator).strip()
            code_lower = indicator.lower().strip()

            if state_name not in state_codes:
                row_num += 1
                continue

            state_id = state_codes[state_name]["id"]
            district_id = state_codes[state_name].get(district_name)

            if not district_id:
                row_num += 1
                continue

            # Handle values
            try:
                if isinstance(data_value, str) and "%" in data_value:
                    processed_value = float(data_value.strip().replace("%", ""))
                elif isinstance(data_value, (int, float)):
                    processed_value = int(data_value) if data_value == int(data_value) else data_value
                elif isinstance(data_value, str):
                    processed_value = int(float(data_value)) if data_value.strip().isdigit() else data_value
                else:
                    processed_value = 0
            except:
                processed_value = 0

            # Init state entry
            if state_id not in states_map:
                states_map[state_id] = {
                    "state_name": state_name,
                    "districts": {}
                }

            # Init district entry
            if district_id not in states_map[state_id]["districts"]:
                states_map[state_id]["districts"][district_id] = {
                    "label": district_name,
                    "type": "category_4",
                    "details": [],
                    "state_led": 0,
                    "district_led": 0
                }

            district_entry = states_map[state_id]["districts"][district_id]

            # ✅ Init district metrics/pie storage
            if district_id not in district_files_map:
                district_files_map[district_id] = {
                    "district_name": district_name,
                    "metrics": [],
                    "pie": []
                }

            # Track for category calculation
            if code_lower == "state led program":
                district_entry["state_led"] = processed_value
            elif code_lower == "district led program":
                district_entry["district_led"] = processed_value

            # ✅ Add to metrics (excluding categories/state/district/community led)
            if code_lower not in excluded_for_metrics:
                district_files_map[district_id]["metrics"].append({
                    "label": indicator.replace("\n", " ").strip(),
                    "value": format_metric_value(data_value)
                })

            # Skip unwanted indicators from details
            if code_lower in excluded_codes:
                # ✅ For pie-chart we still need categories
                if code_lower == "categories":
                    district_files_map[district_id]["pie"].append({
                        "name": str(definition).strip(),
                        "value": processed_value
                    })
                row_num += 1
                continue

            # Add to details
            district_entry["details"].append({
                "value": processed_value,
                "code": indicator
            })

            row_num += 1

        workbook.close()

        # Assign category type for each district
        for state_id, state_data in states_map.items():
            for dist_id, dist_data in state_data["districts"].items():
                state_led = dist_data["state_led"]
                district_led = dist_data["district_led"]
                if state_led > 0 and district_led > 0:
                    dist_data["type"] = "category_1"
                elif state_led > 0:
                    dist_data["type"] = "category_2"
                elif district_led > 0:
                    dist_data["type"] = "category_3"
                else:
                    dist_data["type"] = "category_4"
                # remove helper keys
                dist_data.pop("state_led", None)
                dist_data.pop("district_led", None)

                # Always append "Districts driving improvements" with '-'
                dist_data["details"].insert(1,{
                    "value": "-",
                    "code": "Districts driving improvements"
                })

        # ✅ Save per-district metrics.json & pie-chart.json
        script_dir = os.path.dirname(os.path.abspath(__file__))
        for dist_id, dist_files in district_files_map.items():
            dist_dir = os.path.join(script_dir, "..", "districts", str(dist_id))
            os.makedirs(dist_dir, exist_ok=True)

            metrics_path = os.path.join(dist_dir, "metrics.json")
            with open(metrics_path, "w", encoding="utf-8") as f:
                json.dump({"metrics": dist_files["metrics"]}, f, indent=2, ensure_ascii=False)

            pie_path = os.path.join(dist_dir, "pie-chart.json")
            with open(pie_path, "w", encoding="utf-8") as f:
                json.dump({"data": dist_files["pie"]}, f, indent=2, ensure_ascii=False)

        # ✅ Upload everything (states + districts)
        gcp_access_path = os.path.join(script_dir, '..', 'cloud-scripts', 'gcp_access.py')
        spec = importlib.util.spec_from_file_location('gcp_access', gcp_access_path)
        gcp_access = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(gcp_access)

        # Upload state map.json
        for state_id, state_data in states_map.items():
            json_file_path = os.path.join(script_dir, "..", "states", state_id, "map.json")
            os.makedirs(os.path.dirname(json_file_path), exist_ok=True)

            if os.path.exists(json_file_path):
                with open(json_file_path, "r", encoding="utf-8") as f:
                    existing_json = json.load(f)
            else:
                existing_json = {"result": {}}

            existing_json.setdefault("result", {})
            existing_json["result"]["districts"] = {
                k: v for k, v in state_data["districts"].items()
            }

            with open(json_file_path, "w", encoding="utf-8") as f:
                json.dump(existing_json, f, indent=2, ensure_ascii=False)
            folder_url = gcp_access.upload_file_to_gcs_and_get_directory(
                bucket_name=os.environ.get("BUCKET_NAME"),
                source_file_path=json_file_path,
                destination_blob_name=f"sg-dashboard/states/{state_id}/map.json"
            )

            if folder_url:
                print(f"✅ Uploaded map.json for {state_data['state_name']} to {folder_url}")
            else:
                print(f"❌ Failed to upload map.json for {state_data['state_name']}")

        # ✅ Upload district files
        for dist_id in district_files_map.keys():
            dist_dir = os.path.join(script_dir, "..", "districts", str(dist_id))

            for filename in ["metrics.json", "pie-chart.json"]:
                file_path = os.path.join(dist_dir, filename)
                folder_url = gcp_access.upload_file_to_gcs_and_get_directory(
                    bucket_name=os.environ.get("BUCKET_NAME"),
                    source_file_path=file_path,
                    destination_blob_name=f"sg-dashboard/districts/{dist_id}/{filename}"
                )
                if folder_url:
                    print(f"✅ Uploaded {filename} for district {dist_id} to {folder_url}")
                else:
                    print(f"❌ Failed to upload {filename} for district {dist_id}")

        # Upload state-details-page.json
        state_details_path = os.path.join(script_dir, "..", "pages", "state-details-page.json")
        
        folder_url = gcp_access.upload_file_to_gcs_and_get_directory(
            bucket_name=os.environ.get("BUCKET_NAME"),
            source_file_path=state_details_path,
            destination_blob_name="sg-dashboard/state-details-page.json"
        )

        if folder_url:
            print(f"✅ Uploaded state-details-page.json to {folder_url}")
        else:
            print("❌ Failed to upload state-details-page.json")

    except Exception as e:
        print(f"❌ Error: {str(e)}")

if __name__ == "__main__":
    import sys
    if len(sys.argv) < 2:
        print("Usage: python extract_district_details.py <excel_file>")
    else:
        extract_district_details(sys.argv[1])
