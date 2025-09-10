import openpyxl
import json
import os
from collections import defaultdict
from constants import PAGE_METADATA, TABS_METADATA
import importlib.util

try:
    from state_code_generator import state_code_generator
except ImportError:
    import importlib.util
    import os
    script_dir = os.path.dirname(os.path.abspath(__file__))
    state_gen_path = os.path.join(script_dir, 'state_code_generator.py')

    if os.path.exists(state_gen_path):
        spec = importlib.util.spec_from_file_location('state_code_generator', state_gen_path)
        state_gen_module = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(state_gen_module)
        state_code_generator = state_gen_module.state_code_generator
    else:
        def state_code_generator(excel_file):
            print("state_code_generator function not available")
            return False

def load_state_codes(excel_file):
    try:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        state_codes_file = os.path.join(script_dir, "..", "pages", "state_code_details.json")
        state_code_generator(excel_file)
        if not os.path.exists(state_codes_file):
            return None
        with open(state_codes_file, 'r', encoding='utf-8') as file:
            return json.load(file)
    except Exception:
        return None

def save_and_upload_state_file(script_dir, state_id, filename, data, gcp_access):
    """Save JSON to /states/{id}/filename and upload to GCS."""
    states_dir = os.path.join(script_dir, "..", "states", str(state_id))
    os.makedirs(states_dir, exist_ok=True)

    file_path = os.path.join(states_dir, filename)
    with open(file_path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

    destination_blob_name = f"sg-dashboard/states/{state_id}/{filename}"
    folder_url = gcp_access.upload_file_to_gcs_and_get_directory(
        bucket_name=os.environ.get("BUCKET_NAME"),
        source_file_path=file_path,
        destination_blob_name=destination_blob_name
    )
    print(f"Uploaded {filename} for state {state_id}: {folder_url}")

def update_district_view_indicators(excel_file):
    try:
        state_codes = load_state_codes(excel_file)
        if not state_codes:
            return

        script_dir = os.path.dirname(os.path.abspath(__file__))
        json_file_path = os.path.join(script_dir, "..", "pages", "district-view-indicators.json")

        workbook = openpyxl.load_workbook(excel_file, data_only=False)  # data_only=False to get formatted values

        # --- STEP 1: Extract special indicators from HOME_PAGE tab ---
        try:
            home_page_sheet = workbook[PAGE_METADATA["HOME_PAGE"]]
        except KeyError:
            print(f"Sheet not found: {PAGE_METADATA['HOME_PAGE']}")
            return

        home_expected_headers = TABS_METADATA["HOME_PAGE"]
        home_col_indices = {}
        for cell in home_page_sheet[1]:
            if cell.value and cell.value.strip() in home_expected_headers:
                home_col_indices[cell.value.strip()] = cell.column

        special_keys_lower = ["momentum partners", "nas grade 3"]
        special_indicators_map = {}

        for row in home_page_sheet.iter_rows(min_row=2):
            indicator_name = str(row[home_col_indices["Indicator"] - 1].value or "").strip()
            if indicator_name.lower() in special_keys_lower:
                data_cell = row[home_col_indices["Data"] - 1]
                if indicator_name.lower() == "nas grade 3":
                    # Get the value for NAS Grade 3 and convert to integer string
                    data_val = data_cell.internal_value if data_cell.internal_value else ''
                    if isinstance(data_val, (int, float)):
                        data_val = str(int(data_val))  # Convert to integer and then to string
                    else:
                        data_val = str(data_val)  # Fallback to string if not a number
                else:
                    data_val = data_cell.value
                    if isinstance(data_val, str) and "%" in data_val:
                        data_val = data_val.strip()
                special_indicators_map[indicator_name] = data_val

        # --- STEP 2: Process STATE_DETAILS tab ---
        try:
            sheet = workbook[PAGE_METADATA["STATE_DETAILS"]]
        except KeyError:
            print(f"Sheet not found: {PAGE_METADATA['STATE_DETAILS']}")
            return

        expected_headers = TABS_METADATA["STATE_DETAILS"]
        column_indices = {}
        for cell in sheet[1]:
            if cell.value and cell.value.strip() in expected_headers:
                column_indices[cell.value.strip()] = cell.column

        missing_columns = [col for col in expected_headers if col not in column_indices]
        if missing_columns:
            print(f"Missing required columns: {missing_columns}")
            return

        states_data = {}
        states_mission_data = {}
        overview_aggregates = defaultdict(int)

        # New collectors for per-state files
        state_collectors = {}

        row_num = 2
        while True:
            state_name = sheet.cell(row=row_num, column=column_indices["State Name"]).value
            if not state_name:
                break

            indicator = sheet.cell(row=row_num, column=column_indices["Indicator"]).value or ""
            definition = sheet.cell(row=row_num, column=column_indices["Definition"]).value or ""
            data_value = sheet.cell(row=row_num, column=column_indices["Data"]).value

            indicator = str(indicator).strip()
            definition = str(definition).strip()
            state_name = str(state_name).strip()
            code_lower = indicator.lower().strip()

            if state_name not in state_codes:
                row_num += 1
                continue
            state_code = state_codes[state_name]["id"]

            # normalize value
            try:
                if isinstance(data_value, str) and "%" in data_value:
                    processed_value = data_value.strip()
                elif isinstance(data_value, (int, float)):
                    processed_value = int(data_value) if data_value == int(data_value) else data_value
                elif isinstance(data_value, str) and data_value.strip().isdigit():
                    processed_value = int(data_value.strip())
                else:
                    processed_value = data_value
            except:
                processed_value = 0

            # init collector
            state_collectors.setdefault(state_code, {
                "name": state_name,
                "missions": [],
                "categories": [],
                "map_details": []
            })

            # --- Build per-state collectors ---
            if code_lower in ["state led missions", "district led missions", "community led missions"]:
                identifier_map = {
                    "state led missions": "slm",
                    "district led missions": "dlm",
                    "community led missions": "clm"
                }

                state_collectors[state_code]["missions"].append({
                    "label": indicator,
                    "value": processed_value,
                    "identifier": identifier_map.get(code_lower, "")
                })

            elif "categories" in code_lower:
                state_collectors[state_code]["categories"].append({
                    "name": definition,
                    "value": processed_value
                })

            else:
                state_collectors[state_code]["map_details"].append({
                    "code": indicator,
                    "value": processed_value
                })

            # --- Existing district-view-indicators.json aggregation ---
            if state_code not in states_data:
                states_data[state_code] = {
                    "label": state_name,
                    "type": "category_4",
                    "details": []
                }
                states_mission_data[state_code] = {
                    "state_led_missions": 0,
                    "district_led_missions": 0
                }

            if code_lower == "state led missions":
                states_mission_data[state_code]["state_led_missions"] = processed_value
            elif code_lower == "district led missions":
                states_mission_data[state_code]["district_led_missions"] = processed_value

            excluded_codes = ["categories", "state led missions", "district led missions", "community led missions"]
            if code_lower not in excluded_codes:
                states_data[state_code]["details"].append({
                    "value": processed_value,
                    "code": indicator
                })
                if code_lower not in special_keys_lower and isinstance(processed_value, int):
                    overview_aggregates[indicator] += processed_value

            row_num += 1

        # --- STEP 3: Assign category types for states ---
        for code, data in states_data.items():
            state_led = states_mission_data[code]["state_led_missions"]
            district_led = states_mission_data[code]["district_led_missions"]

            if state_led > 0 and district_led > 0:
                data["type"] = "category_1"
            elif state_led > 0:
                data["type"] = "category_2"
            elif district_led > 0:
                data["type"] = "category_3"
            else:
                data["type"] = "category_4"

        workbook.close()

        # --- STEP 4: Load or create JSON ---
        if os.path.exists(json_file_path):
            with open(json_file_path, 'r', encoding='utf-8') as f:
                district_indicators = json.load(f)
        else:
            district_indicators = {
                "result": {
                    "states": {},
                    "overview": {
                        "label": "india",
                        "type": "category_4",
                        "details": []
                    },
                    "meta": {}
                }
            }

        # --- STEP 5: Prepare overview details ---
        overview_details = [
            {"code": key, "value": value} for key, value in overview_aggregates.items()
        ]

        # Append special indicators from HOME_PAGE
        for special_key_lower in special_keys_lower:
            for orig_key, val in special_indicators_map.items():
                if orig_key.lower() == special_key_lower:
                    overview_details.append({
                        "code": orig_key,
                        "value": val
                    })

        district_indicators["result"]["overview"] = {
            "label": "india",
            "type": "category_4",
            "details": overview_details
        }
        district_indicators["result"]["states"] = states_data

        # --- STEP 6: Save district-view-indicators.json ---
        with open(json_file_path, 'w', encoding='utf-8') as f:
            json.dump(district_indicators, f, indent=2, ensure_ascii=False)

        # Load GCP uploader
        gcp_access_path = os.path.join(script_dir, '..', 'cloud-scripts', 'gcp_access.py')
        spec = importlib.util.spec_from_file_location('gcp_access', gcp_access_path)
        gcp_access = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(gcp_access)
        private_key_path = os.path.join(script_dir, "..", "private-key.json")

        # Upload main file
        folder_url = gcp_access.upload_file_to_gcs_and_get_directory(
            bucket_name=os.environ.get("BUCKET_NAME"),
            source_file_path=json_file_path,
            destination_blob_name="sg-dashboard/district-view-indicators.json"
        )
        print(f"Uploaded district-view-indicators.json: {folder_url}")

        # --- STEP 7: Save & upload per-state files ---
        for state_id, data in state_collectors.items():
            # metrics.json
            metrics = {"metrics": data["missions"]}
            save_and_upload_state_file(script_dir, state_id, "metrics.json", metrics, gcp_access)

            # pie-chart.json
            pie_chart = {"data": data["categories"]}
            save_and_upload_state_file(script_dir, state_id, "pie-chart.json", pie_chart, gcp_access)

            # map.json
            map_json = {
                "result": {
                    "districts": {},
                    "overview": {
                        "label": data["name"].lower(),
                        "type": "category_4",
                        "details": data["map_details"]
                    }
                }
            }
            save_and_upload_state_file(script_dir, state_id, "map.json", map_json, gcp_access)

        print("✅ All files updated & uploaded successfully.")

    except Exception as e:
        print(f"❌ Error: {str(e)}")

if __name__ == "__main__":
    update_district_view_indicators()