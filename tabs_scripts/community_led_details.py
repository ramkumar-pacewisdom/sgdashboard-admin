import openpyxl
import json
import os
import importlib.util

from constants import PAGE_METADATA, TABS_METADATA

def pie_chart_community_led(excel_file):
    try:
        # Get the directory of the current script
        script_dir = os.path.dirname(os.path.abspath(__file__))
        print(script_dir)
        # Define the path to the JSON file
        json_path = os.path.join(script_dir, "..", "pages", "community-led-improvements-page.json")

        # Open the Excel file
        workbook = openpyxl.load_workbook(excel_file, data_only=True)
        try:
            sheet = workbook["Community Led Programs"]
        except KeyError:
            print("Error: Sheet 'Community Led Programs' not found in the Excel file.")
            print(f"Available sheets: {workbook.sheetnames}")
            return

        # Get headers
        headers = [cell.value for cell in sheet[1]]
        cleaned_headers = [str(cell).strip() if cell is not None else '' for cell in headers]
        expected_columns = [
            "Community Engagement",
            "Infrastructure and resources",
            "School structure and practices",
            "Leadership",
            "Pedagogy",
            "Assessment and Evaluation",
        ]

        DISPLAY_NAMES = {
            "Infrastructure and resources": "Infrastructure and Resources",
            "School structure and practices": "School Structure and Practices",
            "Leadership": "Leadership",
            "Pedagogy": "Pedagogy",
            "Assessment and Evaluation": "Assessment and Evaluation",
            "Community Engagement": "Community Engagement"
        }

        if not all(col in cleaned_headers for col in expected_columns):
            print(f"Error: Excel file must contain columns: {expected_columns}")
            print(f"Found: {cleaned_headers}")
            return

        # Calculate sum for each specified column
        data = []
        for col_name in expected_columns:
            col_index = cleaned_headers.index(col_name)
            col_sum = 0
            for row in sheet.iter_rows(min_row=2, max_col=len(headers), values_only=True):
                try:
                    value = row[col_index]
                    if isinstance(value, (int, float)) and value is not None:
                        col_sum += value
                except Exception as e:
                    print(f"Error processing value in column {col_name}: {str(e)}")
                    continue
            # Convert to int if the sum is a whole number
            if isinstance(col_sum, float) and col_sum.is_integer():
                col_sum = int(col_sum)
            data.append({
                'name': DISPLAY_NAMES.get(col_name.strip(), col_name.strip()),
                'value': col_sum
            })

        # Read the existing JSON file
        with open(json_path, 'r', encoding='utf-8') as json_file:
            raw_content = json_file.read()
            json_file.seek(0)
            try:
                json_data = json.load(json_file)
            except json.JSONDecodeError:
                try:
                    json_data = json.loads(raw_content)
                except json.JSONDecodeError:
                    json_data = []

        # Ensure json_data is a list
        if not isinstance(json_data, list):
            json_data = [json_data] if json_data else []

        # Find and update the "pie-chart-community-led" object
        found = False
        for obj in json_data:
            if isinstance(obj, dict) and obj.get('type', '').strip().lower() == 'pie-chart':
                print(data)
                obj['data'] = data
                found = True
                break

        # If not found, append a new object
        if not found:
            json_data.append({'type': 'pie-chart-community-led', 'data': data})

        # Write back to the file
        with open(json_path, 'w', encoding='utf-8') as json_file:
            json.dump(json_data, json_file, indent=2, ensure_ascii=False)

        # Dynamically import gcp_access module and upload file
        gcp_access_path = os.path.join(script_dir, '..', 'cloud-scripts', 'gcp_access.py')
        spec = importlib.util.spec_from_file_location('gcp_access', gcp_access_path)
        gcp_access = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(gcp_access)

        folder_url = gcp_access.upload_file_to_gcs_and_get_directory(
            bucket_name=os.environ.get("BUCKET_NAME"),
            source_file_path=json_path,
            destination_blob_name="sg-dashboard/community-led-improvements-page.json"
        )

        if folder_url:
            print(f"Successfully uploaded and got public folder URL: {folder_url}")
        else:
            print("Failed to upload file to GCS. Check logs for details.")

    except Exception as e:
        print(f"Error: {str(e)}")

def community_led_programs_sum_with_codes(excel_file):
    try:
        # Get the directory of the current script
        script_dir = os.path.dirname(os.path.abspath(__file__))
        print(script_dir)
        # Define the path to the JSON file
        json_path = os.path.join(script_dir, "..", "pages", "community-country-view.json")

        # Open the Excel file
        workbook = openpyxl.load_workbook(excel_file, data_only=True)

        # Load "Community Led Programs" sheet
        try:
            community_sheet = workbook["Community Led Programs"]
        except KeyError:
            print("Error: Sheet 'Community Led Programs' not found in the Excel file.")
            print(f"Available sheets: {workbook.sheetnames}")
            return

        # Load "State_district details" sheet for state codes
        try:
            state_district_sheet = workbook["State_district details"]
        except KeyError:
            print("Error: Sheet 'State_district details' not found in the Excel file.")
            print(f"Available sheets: {workbook.sheetnames}")
            return

        # Get headers for Community Led Programs
        community_headers = [cell.value for cell in community_sheet[1]]
        community_cleaned_headers = [str(cell).strip() if cell is not None else '' for cell in community_headers]
        expected_community_columns = [
            "Name of the State",
            "Name of the District",  # Added district column
            "No. of community leaders engaged",
            "Community led improvements",
            "Challenges shared",
            "Solutions shared"
        ]
        if not all(col in community_cleaned_headers for col in expected_community_columns):
            print(f"Error: Excel file must contain columns in Community Led Programs: {expected_community_columns}")
            print(f"Found: {community_cleaned_headers}")
            return

        # Get headers for State_district details
        state_district_headers = [cell.value for cell in state_district_sheet[1]]
        state_district_cleaned_headers = [str(cell).strip() if cell is not None else '' for cell in state_district_headers]
        expected_state_columns = ["state name", "state code"]
        if not all(col in state_district_cleaned_headers for col in expected_state_columns):
            print(f"Error: Excel file must contain columns in State_district details: {expected_state_columns}")
            print(f"Found: {state_district_cleaned_headers}")
            return

        # Extract state codes from State_district details
        state_codes = {}
        for row in state_district_sheet.iter_rows(min_row=2, max_col=len(state_district_headers), values_only=True):
            try:
                state_name = row[state_district_cleaned_headers.index("state name")] or ''
                state_code = row[state_district_cleaned_headers.index("state code")]
                if state_name and state_code:
                    state_codes[state_name] = str(state_code)
            except Exception as e:
                print(f"Error processing state code row: {str(e)}")
                continue

        # Initialize dictionary to store sums and district counts by state
        state_sums = {}

        # Extract and sum data for specified columns and count unique districts
        for row in community_sheet.iter_rows(min_row=2, max_col=len(community_headers), values_only=True):
            try:
                state_name = row[community_cleaned_headers.index("Name of the State")] or ''
                district_name = row[community_cleaned_headers.index("Name of the District")] or ''
                if not state_name or not district_name:
                    continue

                # Initialize state entry if not exists
                if state_name not in state_sums:
                    state_sums[state_name] = {
                        "No. of community leaders engaged": 0,
                        "Community led improvements": 0,
                        "Challenges shared": 0,
                        "Solutions shared": 0,
                        "Districts activated": set()  # Use a set to store unique district names
                    }

                # Add district to the set
                state_sums[state_name]["Districts activated"].add(district_name)

                # Sum values for specified columns
                for col_name in expected_community_columns[2:]:  # Skip State Name and District
                    col_index = community_cleaned_headers.index(col_name)
                    value = row[col_index]
                    if isinstance(value, (int, float)) and value is not None:
                        state_sums[state_name][col_name] += value

            except Exception as e:
                print(f"Error processing row in Community Led Programs: {str(e)}")
                continue

        # Format data as object of objects with state code as key
        states_data = {
            state_codes.get(state, "unknown"): {
                "id": state_codes.get(state, "unknown"),
                "label": state,
                "type": "category_1",
                "details": [
                    {"code": col_name, "value": int(val) if isinstance(val, float) and val.is_integer() else val}
                    for col_name, val in sums.items() if col_name != "Districts activated"  # Exclude Districts Activated temporarily
                ] + [{"code": "Districts activated", "value": len(sums["Districts activated"])}]  # Add district count
            }
            for state, sums in state_sums.items()
        }

        # Create the final data structure
        data = {
            "result": {
                "states": states_data
            }
        }

        # Read the existing JSON file
        try:
            with open(json_path, 'r', encoding='utf-8') as json_file:
                raw_content = json_file.read()
                json_file.seek(0)
                try:
                    json_data = json.load(json_file)
                except json.JSONDecodeError:
                    try:
                        json_data = json.loads(raw_content)
                    except json.JSONDecodeError:
                        json_data = {}
        except FileNotFoundError:
            json_data = {}

        # Ensure json_data is a dictionary with a "result" key
        if not isinstance(json_data, dict):
            json_data = {"result": {}}
        if "result" not in json_data:
            json_data["result"] = {}
        if "states" not in json_data["result"]:
            json_data["result"]["states"] = {}

        # Update the states data
        print(states_data)
        json_data["result"]["states"].update(states_data)

        # Write back to the file
        with open(json_path, 'w', encoding='utf-8') as json_file:
            json.dump(json_data, json_file, indent=2, ensure_ascii=False)

        # Dynamically import gcp_access module and upload file
        gcp_access_path = os.path.join(script_dir, '..', 'cloud-scripts', 'gcp_access.py')
        spec = importlib.util.spec_from_file_location('gcp_access', gcp_access_path)
        gcp_access = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(gcp_access)

        folder_url = gcp_access.upload_file_to_gcs_and_get_directory(
            bucket_name=os.environ.get("BUCKET_NAME"),
            source_file_path=json_path,
            destination_blob_name="sg-dashboard/community-country-view.json"
        )

        if folder_url:
            print(f"Successfully uploaded and got public folder URL: {folder_url}")
            updateOverviewValues()
        else:
            print("Failed to upload file to GCS. Check logs for details.")

    except Exception as e:
        print(f"Error: {str(e)}")

def updateOverviewValues():
    # Get the directory of the script
    script_dir = os.path.dirname(os.path.abspath(__file__))

    # Construct the JSON file path
    json_path = os.path.join(script_dir, "..", "pages", "community-country-view.json")

    # Read JSON data from file with detailed error handling
    try:
        with open(json_path, 'r', encoding='utf-8') as f:
            file_content = f.read()
            if not file_content.strip():
                print(f"Error: The file {json_path} is empty.")
                exit(1)
            data = json.loads(file_content)
    except FileNotFoundError:
        print(f"Error: File not found at {json_path}")
        exit(1)
    except UnicodeDecodeError as e:
        print(f"Error: Encoding issue in {json_path}. Ensure the file is UTF-8 encoded.")
        print(f"Details: {str(e)}")
        exit(1)
    except JSONDecodeError as e:
        print(f"Error: Invalid JSON format in {json_path}")
        print(f"Details: {str(e)}")
        print(f"Line: {e.lineno}, Column: {e.colno}")
        print(f"Content near error: {e.doc[max(0, e.pos-20):e.pos+20]}")
        exit(1)
    except Exception as e:
        print(f"Unexpected error while reading {json_path}: {str(e)}")
        exit(1)

    # Initialize a dictionary to store the sums for each code dynamically
    code_sums = {}

    # Navigate each state and sum the values for each code in details
    try:
        for state_id, state_data in data['result']['states'].items():
            for detail in state_data['details']:
                code = detail['code']
                value = detail['value']
                # Initialize the code in code_sums if not present
                if code not in code_sums:
                    code_sums[code] = 0
                code_sums[code] += value
    except KeyError as e:
        print(f"Error: Missing expected key in JSON structure: {str(e)}")
        exit(1)

    # Ensure all codes from code_sums exist in overview details
    try:
        overview_details = data['result']['overview']['details']
        existing_codes = {detail['code'] for detail in overview_details}
        
        # Add missing codes to overview details
        for code in code_sums:
            if code not in existing_codes:
                print(f"Adding missing code '{code}' to overview details")
                overview_details.append({"code": code, "value": 0})
        
        # Update the overview details with the summed values
        for detail in overview_details:
            code = detail['code']
            if code in code_sums:
                detail['value'] = code_sums[code]
            else:
                print(f"Warning: Code '{code}' in overview not found in states")
    except KeyError as e:
        print(f"Error: Missing expected key in overview structure: {str(e)}")
        exit(1)

    # Convert the updated data to a JSON string with indentation
    updated_json = json.dumps(data, indent=2)

    # Print the updated JSON
    print(updated_json)

    # Save the updated JSON to a new file in the same directory
    output_path = os.path.join(script_dir, "..", "pages", "community-country-view.json")
    try:
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2)
        gcp_access_path = os.path.join(script_dir, '..', 'cloud-scripts', 'gcp_access.py')
        spec = importlib.util.spec_from_file_location('gcp_access', gcp_access_path)
        gcp_access = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(gcp_access)
        folder_url = gcp_access.upload_file_to_gcs_and_get_directory(
            bucket_name=os.environ.get("BUCKET_NAME"),
            source_file_path=json_path,
            destination_blob_name="sg-dashboard/community-country-view.json"
        )
        print(f"Updated JSON saved to {output_path}")
    except Exception as e:
        print(f"Error: Failed to write to {output_path}: {str(e)}")
        exit(1)




                