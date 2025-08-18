import openpyxl
import json
import os
from constants import PAGE_METADATA, TABS_METADATA
import importlib.util


def state_code_generator(excel_file):
    try:
        # Get the script directory and file path
        script_dir = os.path.dirname(os.path.abspath(__file__))
        
        # # Check if Excel file exists
        # if not os.path.exists(file_path):
        #     print(f"Error: Excel file not found at {file_path}")
        #     return
        
        # Load workbook
        workbook = openpyxl.load_workbook(excel_file, data_only=True)
        print(f"WORKBOOK IN STATEGEN {workbook.sheetnames}")
        try:
            sheet = workbook[PAGE_METADATA["STATE_DISTRICT_DETAILS"]]
        except KeyError:
            print("Error: Sheet not found in the Excel file.")
            print(f"Available sheets: {workbook.sheetnames}")
            return
        
        # Get column headers from TABS_METADATA
        expected_headers = TABS_METADATA["STATE_DISTRICT_DETAILS"]
        print(f"Expected headers: {expected_headers}")
        
        # Read headers from the first row to find column indices
        header_row = 1
        column_indices = {}
        
        for cell in sheet[header_row]:
            if cell.value and cell.value.lower().strip() in [h.lower() for h in expected_headers]:
                for header in expected_headers:
                    if cell.value.lower().strip() == header.lower():
                        column_indices[header] = cell.column
                        break
        
        print(f"Found column indices: {column_indices}")
        
        # Verify we have the required columns
        required_columns = TABS_METADATA["STATE_DISTRICT_DETAILS"]
        missing_columns = [col for col in required_columns if col not in column_indices]
        
        if missing_columns:
            print(f"Error: Missing required columns: {missing_columns}")
            return
        
        # Final JSON structure
        json_data = {}
        
        # Start from row 2 (assuming row 1 has headers)
        row_num = 2
        processed_count = 0
        
        while True:
            state_name_cell = sheet.cell(row=row_num, column=column_indices["state name"])
            district_name_cell = sheet.cell(row=row_num, column=column_indices["district name"])
            state_code_cell = sheet.cell(row=row_num, column=column_indices["state code"])
            district_code_cell = sheet.cell(row=row_num, column=column_indices["district code"])
            
            # Check if we've reached the end of data
            if not state_name_cell.value:
                break
            
            state_name = str(state_name_cell.value).strip()
            district_name = str(district_name_cell.value).strip() if district_name_cell.value else None
            state_code = str(state_code_cell.value).strip() if state_code_cell.value else None
            district_code = str(district_code_cell.value).strip() if district_code_cell.value else None
            
            if state_name not in json_data:
                json_data[state_name] = {"id": state_code}

            # Add district info if available
            if district_name and district_code:
                json_data[state_name][district_name] = district_code
            
            processed_count += 1
            row_num += 1
        
        print(f"Processed {processed_count} rows, found {len(json_data)} states")
        
        # Define output file path
        output_file = os.path.join(script_dir, "..", "pages", "state_code_details.json")
        india_json_file = os.path.join(script_dir, "..", "pages", "india.json")
        
        # Create or overwrite the JSON file
        try:
            with open(output_file, 'w', encoding='utf-8') as json_file:
                json.dump(json_data, json_file, indent=4, ensure_ascii=False)

            gcp_access_path = os.path.join(script_dir, '..', 'cloud-scripts', 'gcp_access.py')
            spec = importlib.util.spec_from_file_location('gcp_access', gcp_access_path)
            gcp_access = importlib.util.module_from_spec(spec)
            spec.loader.exec_module(gcp_access)

            private_key_path = os.path.join(script_dir, "..", "private-key.json")

            folder_url = gcp_access.upload_file_to_gcs_and_get_directory(
                bucket_name=os.environ.get("BUCKET_NAME"),
                source_file_path=output_file,
                destination_blob_name="sg-dashboard/state_code_details.json"
            )

            if folder_url:
                print(f"Successfully uploaded and got public folder URL state code gen: {folder_url}")
            else:
                print("Failed to upload file to GCS. Check logs for details.")

            folder_url_for_india_json = gcp_access.upload_file_to_gcs_and_get_directory(
                bucket_name=os.environ.get("BUCKET_NAME"),
                source_file_path=india_json_file,
                destination_blob_name="sg-dashboard/india.json"
            )

            if folder_url_for_india_json:
                print(f"Successfully uploaded and got public folder URL: {folder_url_for_india_json}")
            else:
                print("Failed to upload file to GCS. Check logs for details.")

            
            print(f"Successfully created {output_file}")
                    
        except Exception as e:
            print(f"Error writing to JSON file: {str(e)}")
        
        # Close workbook
        workbook.close()
        
    except FileNotFoundError:
        print(f"Error: Excel file not found at {file_path}")
    except Exception as e:
        print(f"Error: {str(e)}")


if __name__ == "__main__":
    state_code_generator(excel_file)