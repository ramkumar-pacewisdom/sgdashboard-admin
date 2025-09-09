import openpyxl
import json
import os
import importlib.util
import re
import requests

from constants import PAGE_METADATA,TABS_METADATA

def convert_drive_link_to_direct_url(link):
    if not isinstance(link, str):
        return ''
    match = re.search(r"/d/([a-zA-Z0-9_-]+)", link)
    if not match:
        match = re.search(r"id=([a-zA-Z0-9_-]+)", link)
    if match:
        file_id = match.group(1)
        return f"https://drive.google.com/uc?export=view&id={file_id}"
    return link.strip()

def download_image(file_id, save_path):
    url = f"https://drive.google.com/uc?export=download&id={file_id}"
    try:
        response = requests.get(url)
        if response.status_code == 200:
            with open(save_path, 'wb') as f:
                f.write(response.content)
            print(f"✅ Downloaded image to {save_path}")
            return True
        else:
            print(f"❌ Failed to download image for file ID {file_id}")
            return False
    except Exception as e:
        print(f"❌ Exception downloading image: {e}")
        return False

def key_progress_indicators(excel_file):
    try:
        # Get the directory of the current script
        script_dir = os.path.dirname(os.path.abspath(__file__))
        print(script_dir)
        # Define the path to the JSON file
        json_path = os.path.join(script_dir, "..", "pages", "landing-page.json")
        images_dir = os.path.join(script_dir, "temp_downloads")
        os.makedirs(images_dir, exist_ok=True)

        # Open the Excel file
        workbook = openpyxl.load_workbook(excel_file, data_only=False)  # data_only=False to get formatted values
        try:
            sheet = workbook[PAGE_METADATA["HOME_PAGE"]]
        except KeyError:
            print("Error: Sheet 'Data on homepage' not found in the Excel file.")
            print(f"Available sheets: {workbook.sheetnames}")
            return

        # Get headers
        headers = [cell.value for cell in sheet[1]]
        cleaned_headers = [str(cell).strip() if cell is not None else '' for cell in headers]
        expected_columns = TABS_METADATA["HOME_PAGE"]
        if not all(col in cleaned_headers for col in expected_columns):
            print(f"Error: Excel file must contain columns: {expected_columns}")
            print(f"Found: {cleaned_headers}")
            return

        # Extract data rows
        data = []
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):  # Use cell objects, not values_only
            try:
                raw_name = row[cleaned_headers.index(TABS_METADATA["HOME_PAGE"][0])].value
                if not raw_name or not str(raw_name).strip():
                    continue

                raw_src = row[cleaned_headers.index(TABS_METADATA["HOME_PAGE"][3])].value or ''
                logo_url = convert_drive_link_to_direct_url(raw_src)

                file_match = re.search(r"id=([a-zA-Z0-9_-]+)", logo_url)
                if not file_match:
                    file_match = re.search(r"/d/([a-zA-Z0-9_-]+)", raw_src)
                file_id = file_match.group(1) if file_match else ''

                name_clean = str(raw_name).strip().lower()
                name_clean = re.sub(r'[^a-z0-9_-]', '', name_clean.replace(" ", "_"))
                local_filename = f"{name_clean}.svg"
                local_path = os.path.join(images_dir, local_filename)

                if file_id:
                    if download_image(file_id, local_path):
                         gcp_access_path = os.path.join(script_dir, '..', 'cloud-scripts', 'gcp_access.py')
                         spec = importlib.util.spec_from_file_location('gcp_access', gcp_access_path)
                         gcp_access = importlib.util.module_from_spec(spec)
                         spec.loader.exec_module(gcp_access)

                         folder_url = gcp_access.upload_file_to_gcs_and_get_directory(
                            bucket_name=os.environ.get("BUCKET_NAME"),
                            source_file_path=local_path,
                            destination_blob_name = f"sg-dashboard/partners/{local_filename}"
                         )

                         if folder_url:
                            os.remove(local_path)
                            final_src = f"{folder_url.rstrip('/')}/{local_filename}"
                            print(f"Successfully uploaded and got public folder URL: {folder_url}, {final_src}")
                         else:
                            print("Failed to upload file to GCS. Check logs for details.")

                # Get the formatted value for the 'value' column
                value_cell = row[cleaned_headers.index(TABS_METADATA["HOME_PAGE"][2])]
                row_data = {
                    'label': raw_name or '',
                    'value': value_cell.value or '',  # Start with raw value
                    'icon': final_src or ''
                }
                # For 'NAS Grade 3', get the formatted text (e.g., "59%")
                if row_data['label'] == 'NAS Grade 3':
                    row_data['value'] = value_cell.internal_value if value_cell.internal_value else ''
                    if value_cell.number_format and '%' in value_cell.number_format:
                        # If the cell is formatted as percentage, get the display value as an integer
                        row_data['value'] = f"{int(value_cell.value * 100)}%"
                    else:
                        # Ensure whole numbers are stored as integers
                        row_data['value'] = str(int(float(row_data['value']))) if row_data['value'] else ''
                # Convert float to int if it's a whole number for other labels
                elif isinstance(row_data['value'], float) and row_data['value'].is_integer():
                    row_data['value'] = int(row_data['value'])
                
                data.append(row_data)
            except Exception as e:
                print(f"Error processing row {row_idx}: {str(e)}")
                continue

        # Read the existing JSON file
        with open(json_path, 'r', encoding='utf-8') as json_file:
            raw_content = json_file.read()
            json_file.seek(0)
            try:
                json_data = json.load(json_file)
            except json.JSONDecodeError:
                try:
                    json_data = json.loads(raw_content)
                except json.JSONDecodeError:
                    json_data = []

        # Ensure json_data is a list
        if not isinstance(json_data, list):
            json_data = [json_data] if json_data else []

        # Find and update the "data-indicators" object
        found = False
        for obj in json_data:
            if isinstance(obj, dict) and obj.get('type', '').strip().lower() == 'data-indicators':
                obj['indicators'] = data
                found = True
                break

        # If not found, append (this is a fallback, but your JSON should already have it)
        if not found:
            json_data.append({'type': 'data-indicators', 'indicators': data})

        # Write back to the file
        with open(json_path, 'w', encoding='utf-8') as json_file:
            json.dump(json_data, json_file, indent=2, ensure_ascii=False)

        # Dynamically import gcp_access module and upload file
        gcp_access_path = os.path.join(script_dir, '..', 'cloud-scripts', 'gcp_access.py')
        spec = importlib.util.spec_from_file_location('gcp_access', gcp_access_path)
        gcp_access = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(gcp_access)

        private_key_path = os.path.join(script_dir, "..", "private-key.json")

        folder_url = gcp_access.upload_file_to_gcs_and_get_directory(
            bucket_name=os.environ.get("BUCKET_NAME"),
            source_file_path=json_path,
            destination_blob_name="sg-dashboard/landing-page.json"
        )

        if folder_url:
            print(f"Successfully uploaded and got public folder URL: {folder_url}")
        else:
            print("Failed to upload file to GCS. Check logs for details.")

    except Exception as e:
        print(f"Error: {str(e)}")