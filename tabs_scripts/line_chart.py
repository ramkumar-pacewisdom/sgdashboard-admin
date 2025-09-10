from io import BytesIO
import openpyxl
import json
import os
import importlib.util

from constants import PAGE_METADATA

def extract_micro_improvements(excel_file):
    script_dir = os.path.dirname(os.path.abspath(__file__))
    print(script_dir)
    # Define the path to the JSON file
    json_path = os.path.join(script_dir, "..", "pages", "dashboard.json")

    # Open the Excel file
    workbook = openpyxl.load_workbook(excel_file, data_only=True)
    try:
        sheet = workbook["Micro improvements progress"]
    except KeyError:
        print("Error: Sheet 'Micro improvements progress' not found in the Excel file.")
        print(f"Available sheets: {workbook.sheetnames}")
        return

    print(workbook, sheet)

    # Initialize dictionaries to store sums for each year
    sums_2024 = {'Q1': 0, 'Q2': 0, 'Q3': 0, 'Q4': 0}
    sums_2025 = {'Q1': 0, 'Q2': 0, 'Q3': 0, 'Q4': 0}
    valid_quarters_2024 = {'Q1': False, 'Q2': False, 'Q3': False, 'Q4': False}
    valid_quarters_2025 = {'Q1': False, 'Q2': False, 'Q3': False, 'Q4': False}

    # Iterate through rows, starting from row 2 to skip headers
    for row in sheet.iter_rows(min_row=2, max_col=7, values_only=True):
        print("function called inside", row)
        district_name = row[1]  # District column (B)
        year = row[2]  # Year column (C)
        q1, q2, q3, q4 = row[3:7]  # Q1, Q2, Q3, Q4 columns (D-G)
        print("data", year, q1, q2, q3, q4)

        # Skip rows with district data
        if district_name:
            continue

        # Skip rows where Q1, Q2, Q3, and Q4 are all None or empty
        if not any(x is not None for x in [q1, q2, q3, q4]):
            continue

        # Add values to sums only if they are not None
        if year == 2024:
            if q1 is not None:
                sums_2024['Q1'] += float(q1)
                valid_quarters_2024['Q1'] = True
            if q2 is not None:
                sums_2024['Q2'] += float(q2)
                valid_quarters_2024['Q2'] = True
            if q3 is not None:
                sums_2024['Q3'] += float(q3)
                valid_quarters_2024['Q3'] = True
            if q4 is not None:
                sums_2024['Q4'] += float(q4)
                valid_quarters_2024['Q4'] = True
        elif year == 2025:
            if q1 is not None:
                sums_2025['Q1'] += float(q1)
                valid_quarters_2025['Q1'] = True
            if q2 is not None:
                sums_2025['Q2'] += float(q2)
                valid_quarters_2025['Q2'] = True
            if q3 is not None:
                sums_2025['Q3'] += float(q3)
                valid_quarters_2025['Q3'] = True
            if q4 is not None:
                sums_2025['Q4'] += float(q4)
                valid_quarters_2025['Q4'] = True

    # Format the result, including only quarters with valid (non-None) data
    result = []
    for year, sums, valid_quarters in [(2024, sums_2024, valid_quarters_2024), (2025, sums_2025, valid_quarters_2025)]:
        data = []
        for q in ['Q1', 'Q2', 'Q3', 'Q4']:
            if valid_quarters[q]:
                data.append(sums[q])
        if data:  # Include year only if there is valid data
            result.append({
                "year": year,
                "data": data
            })
    
    print(result)

    try:
        with open(json_path, 'r') as file:
            dashboard_data = json.load(file)

        # Find and update the object with type "line-chart"
        for item in dashboard_data:
            if item.get('type') == 'line-chart':
                item['data'] = result

        # Write the updated data back to dashboard.json
        with open(json_path, 'w') as file:
            json.dump(dashboard_data, file, indent=2)

        print(f"Updated dashboard.json with new line-chart data: {json.dumps(result, indent=2)}")
        extract_state_line_chart(excel_file)
        return json.dumps(dashboard_data, indent=2)

    except FileNotFoundError:
        print(f"Error: dashboard.json not found at {json_path}")
        return json.dumps(result, indent=2)
    except Exception as e:
        print(f"Error updating dashboard.json: {str(e)}")
        return json.dumps(result, indent=2)

def load_state_codes():
    """Load state and district codes from state_code_details.json."""
    script_dir = os.path.dirname(os.path.abspath(__file__))
    state_codes_file = os.path.join(script_dir, "..", "pages", "state_code_details.json")
    if not os.path.exists(state_codes_file):
        print("❌ state_code_details.json not found.")
        return None
    with open(state_codes_file, "r", encoding="utf-8") as f:
        return json.load(f)

def extract_district_line_chart(excel_file):
    """Extract data from 'Micro improvements progress' sheet and generate line-chart.json for each district."""
    try:
        # Load state codes
        state_codes = load_state_codes()
        if not state_codes:
            return

        # Load Excel file
        workbook = openpyxl.load_workbook(excel_file, data_only=True)
        try:
            sheet = workbook["Micro improvements progress"]
        except KeyError:
            print("Error: Sheet 'Micro improvements progress' not found in the Excel file.")
            print(f"Available sheets: {workbook.sheetnames}")
            return

        # Initialize district-level containers for line chart data
        district_files_map = {}

        # Iterate through rows, starting from row 2 to skip headers
        row_num = 2
        while True:
            state_name = sheet.cell(row=row_num, column=1).value
            district_name = sheet.cell(row=row_num, column=2).value
            year = sheet.cell(row=row_num, column=3).value
            q1 = sheet.cell(row=row_num, column=4).value
            q2 = sheet.cell(row=row_num, column=5).value
            q3 = sheet.cell(row=row_num, column=6).value
            q4 = sheet.cell(row=row_num, column=7).value

            if not state_name and not district_name:
                break

            state_name = str(state_name).strip() if state_name else ""
            district_name = str(district_name).strip() if district_name else ""

            # Skip if state_name is not in state_codes
            if state_name not in state_codes:
                print(f"⚠️ State '{state_name}' not found in state_code_details.json, skipping row {row_num}")
                row_num += 1
                continue

            state_id = state_codes[state_name]["id"]
            district_id = state_codes[state_name].get(district_name)

            # Skip if district_id is not found
            if not district_id:
                print(f"⚠️ District '{district_name}' not found for state '{state_name}' in state_code_details.json, skipping row {row_num}")
                row_num += 1
                continue

            # Initialize district entry if not exists
            if district_id not in district_files_map:
                district_files_map[district_id] = {
                    "district_name": district_name,
                    "line_chart": {
                        2024: {'Q1': 0, 'Q2': 0, 'Q3': 0, 'Q4': 0, 'valid_Q1': False, 'valid_Q2': False, 'valid_Q3': False, 'valid_Q4': False},
                        2025: {'Q1': 0, 'Q2': 0, 'Q3': 0, 'Q4': 0, 'valid_Q1': False, 'valid_Q2': False, 'valid_Q3': False, 'valid_Q4': False}
                    }
                }

            # Add values to sums only if they are not None
            if year == 2024:
                if q1 is not None:
                    district_files_map[district_id]["line_chart"][2024]['Q1'] += float(q1)
                    district_files_map[district_id]["line_chart"][2024]['valid_Q1'] = True
                if q2 is not None:
                    district_files_map[district_id]["line_chart"][2024]['Q2'] += float(q2)
                    district_files_map[district_id]["line_chart"][2024]['valid_Q2'] = True
                if q3 is not None:
                    district_files_map[district_id]["line_chart"][2024]['Q3'] += float(q3)
                    district_files_map[district_id]["line_chart"][2024]['valid_Q3'] = True
                if q4 is not None:
                    district_files_map[district_id]["line_chart"][2024]['Q4'] += float(q4)
                    district_files_map[district_id]["line_chart"][2024]['valid_Q4'] = True
            elif year == 2025:
                if q1 is not None:
                    district_files_map[district_id]["line_chart"][2025]['Q1'] += float(q1)
                    district_files_map[district_id]["line_chart"][2025]['valid_Q1'] = True
                if q2 is not None:
                    district_files_map[district_id]["line_chart"][2025]['Q2'] += float(q2)
                    district_files_map[district_id]["line_chart"][2025]['valid_Q2'] = True
                if q3 is not None:
                    district_files_map[district_id]["line_chart"][2025]['Q3'] += float(q3)
                    district_files_map[district_id]["line_chart"][2025]['valid_Q3'] = True
                if q4 is not None:
                    district_files_map[district_id]["line_chart"][2025]['Q4'] += float(q4)
                    district_files_map[district_id]["line_chart"][2025]['valid_Q4'] = True

            row_num += 1

        workbook.close()

        # Save line-chart.json for each district
        script_dir = os.path.dirname(os.path.abspath(__file__))
        for dist_id, dist_data in district_files_map.items():
            dist_dir = os.path.join(script_dir, "..", "districts", str(dist_id))
            os.makedirs(dist_dir, exist_ok=True)

            # Format line chart data, including only quarters with valid data
            line_chart_data = []
            for year in [2024, 2025]:
                data = []
                for q in ['Q1', 'Q2', 'Q3', 'Q4']:
                    if dist_data["line_chart"][year][f'valid_{q}']:
                        data.append(dist_data["line_chart"][year][q])
                if data:  # Include year only if there is valid data
                    line_chart_data.append({
                        "year": year,
                        "data": data
                    })

            # Save line-chart.json
            line_chart_path = os.path.join(dist_dir, "line-chart.json")
            with open(line_chart_path, "w", encoding="utf-8") as f:
                json.dump({"data": line_chart_data}, f, indent=2, ensure_ascii=False)

            print(f"✅ Generated line-chart.json for district {dist_id} ({dist_data['district_name']})")

            # Upload district line-chart.json files to GCP
            gcp_access_path = os.path.join(script_dir, '..', 'cloud-scripts', 'gcp_access.py')
            spec = importlib.util.spec_from_file_location('gcp_access', gcp_access_path)
            gcp_access = importlib.util.module_from_spec(spec)
            spec.loader.exec_module(gcp_access)

            file_path = os.path.join(dist_dir, "line-chart.json")
            folder_url = gcp_access.upload_file_to_gcs_and_get_directory(
                bucket_name=os.environ.get("BUCKET_NAME"),
                source_file_path=file_path,
                destination_blob_name=f"sg-dashboard/districts/{dist_id}/line-chart.json"
            )
            if folder_url:
                print(f"✅ Uploaded line-chart.json for district {dist_id} to {folder_url}")
            else:
                print(f"❌ Failed to upload line-chart.json for district {dist_id}")

    except Exception as e:
        print(f"❌ Error: {str(e)}")

def load_state_codes():
    """Load state codes from state_code_details.json."""
    try:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        state_codes_file = os.path.join(script_dir, "..", "pages", "state_code_details.json")
        if not os.path.exists(state_codes_file):
            print("❌ state_code_details.json not found.")
            return None
        with open(state_codes_file, 'r', encoding='utf-8') as file:
            return json.load(file)
    except Exception as e:
        print(f"❌ Error loading state codes: {str(e)}")
        return None

def save_and_upload_state_file(script_dir, state_id, filename, data, gcp_access):
    """Save JSON to /states/{id}/filename and upload to GCS."""
    states_dir = os.path.join(script_dir, "..", "states", str(state_id))
    os.makedirs(states_dir, exist_ok=True)

    file_path = os.path.join(states_dir, filename)
    with open(file_path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

    destination_blob_name = f"sg-dashboard/states/{state_id}/{filename}"
    folder_url = gcp_access.upload_file_to_gcs_and_get_directory(
        bucket_name=os.environ.get("BUCKET_NAME"),
        source_file_path=file_path,
        destination_blob_name=destination_blob_name
    )
    if folder_url:
        print(f"✅ Uploaded {filename} for state {state_id}: {folder_url}")
    else:
        print(f"❌ Failed to upload {filename} for state {state_id}")

def extract_state_line_chart(excel_file):
    """Extract data from 'Micro improvements progress' sheet and generate line-chart.json for each state, excluding district data."""
    try:
        # Load state codes
        state_codes = load_state_codes()
        if not state_codes:
            return

        # Load Excel file
        workbook = openpyxl.load_workbook(excel_file, data_only=True)
        try:
            sheet = workbook["Micro improvements progress"]
        except KeyError:
            print("Error: Sheet 'Micro improvements progress' not found in the Excel file.")
            print(f"Available sheets: {workbook.sheetnames}")
            return

        # Initialize state-level containers for line chart data
        state_line_chart_map = {}

        # Iterate through rows, starting from row 2 to skip headers
        row_num = 2
        while True:
            state_name = sheet.cell(row=row_num, column=1).value
            district_name = sheet.cell(row=row_num, column=2).value
            year = sheet.cell(row=row_num, column=3).value
            q1 = sheet.cell(row=row_num, column=4).value
            q2 = sheet.cell(row=row_num, column=5).value
            q3 = sheet.cell(row=row_num, column=6).value
            q4 = sheet.cell(row=row_num, column=7).value

            if not state_name:
                break

            # Skip rows with district data
            if district_name:
                row_num += 1
                continue

            state_name = str(state_name).strip() if state_name else ""
            if state_name not in state_codes:
                print(f"⚠️ State '{state_name}' not found in state_code_details.json, skipping row {row_num}")
                row_num += 1
                continue

            state_id = state_codes[state_name]["id"]

            # Initialize state entry if not exists
            if state_id not in state_line_chart_map:
                state_line_chart_map[state_id] = {
                    "state_name": state_name,
                    "line_chart": {
                        2024: {'Q1': 0, 'Q2': 0, 'Q3': 0, 'Q4': 0, 'valid_Q1': False, 'valid_Q2': False, 'valid_Q3': False, 'valid_Q4': False},
                        2025: {'Q1': 0, 'Q2': 0, 'Q3': 0, 'Q4': 0, 'valid_Q1': False, 'valid_Q2': False, 'valid_Q3': False, 'valid_Q4': False}
                    }
                }

            # Add values to sums only if they are not None
            if year == 2024:
                if q1 is not None:
                    state_line_chart_map[state_id]["line_chart"][2024]['Q1'] += float(q1)
                    state_line_chart_map[state_id]["line_chart"][2024]['valid_Q1'] = True
                if q2 is not None:
                    state_line_chart_map[state_id]["line_chart"][2024]['Q2'] += float(q2)
                    state_line_chart_map[state_id]["line_chart"][2024]['valid_Q2'] = True
                if q3 is not None:
                    state_line_chart_map[state_id]["line_chart"][2024]['Q3'] += float(q3)
                    state_line_chart_map[state_id]["line_chart"][2024]['valid_Q3'] = True
                if q4 is not None:
                    state_line_chart_map[state_id]["line_chart"][2024]['Q4'] += float(q4)
                    state_line_chart_map[state_id]["line_chart"][2024]['valid_Q4'] = True
            elif year == 2025:
                if q1 is not None:
                    state_line_chart_map[state_id]["line_chart"][2025]['Q1'] += float(q1)
                    state_line_chart_map[state_id]["line_chart"][2025]['valid_Q1'] = True
                if q2 is not None:
                    state_line_chart_map[state_id]["line_chart"][2025]['Q2'] += float(q2)
                    state_line_chart_map[state_id]["line_chart"][2025]['valid_Q2'] = True
                if q3 is not None:
                    state_line_chart_map[state_id]["line_chart"][2025]['Q3'] += float(q3)
                    state_line_chart_map[state_id]["line_chart"][2025]['valid_Q3'] = True
                if q4 is not None:
                    state_line_chart_map[state_id]["line_chart"][2025]['Q4'] += float(q4)
                    state_line_chart_map[state_id]["line_chart"][2025]['valid_Q4'] = True

            row_num += 1

        workbook.close()

        # Load GCP uploader
        script_dir = os.path.dirname(os.path.abspath(__file__))
        gcp_access_path = os.path.join(script_dir, '..', 'cloud-scripts', 'gcp_access.py')
        spec = importlib.util.spec_from_file_location('gcp_access', gcp_access_path)
        gcp_access = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(gcp_access)

        # Save & upload line-chart.json for each state
        for state_id, state_data in state_line_chart_map.items():
            line_chart_data = {
                "data": []
            }
            for year in [2024, 2025]:
                data = []
                for q in ['Q1', 'Q2', 'Q3', 'Q4']:
                    if state_data["line_chart"][year][f'valid_{q}']:
                        data.append(state_data["line_chart"][year][q])
                if data:  # Include year only if there is valid data
                    line_chart_data["data"].append({
                        "year": year,
                        "data": data
                    })
            if line_chart_data["data"]:  # Save only if there is data
                save_and_upload_state_file(script_dir, state_id, "line-chart.json", line_chart_data, gcp_access)

        print("✅ All line-chart.json files generated & uploaded successfully.")
        extract_district_line_chart(excel_file)

    except Exception as e:
        print(f"❌ Error: {str(e)}")

if __name__ == "__main__":
    import sys
    if len(sys.argv) < 2:
        print("Usage: python extract_line_charts.py <excel_file>")
    else:
        extract_state_line_chart(sys.argv[1])