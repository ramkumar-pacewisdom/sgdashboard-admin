import openpyxl
import json
import os
import time
from geopy.geocoders import Nominatim
from constants import PAGE_METADATA, TABS_METADATA
import importlib.util

# Initialize geolocator
geolocator = Nominatim(user_agent="sgdashboard_network_mapper_v1.0")
location_cache = {}


# Geocode helper (first checks partner data)
def get_coordinates(state, country, partner_id=None, partners_data=None):
    # 1. Check partner's own coords if data is provided
    if partner_id and state and partners_data:
      for p in partners_data:
         if (
            p.get("partner_id", "").lower() == str(partner_id).lower()
            and p.get("state", "").lower() == str(state).lower()
         ):
            coords = p.get("coords")
            if isinstance(coords, list) and len(coords) == 2:
                return coords

    # 2. If no partner coords, fall back to geocoding
    queries = []
    if state and country:
        queries.append(f"{state}, {country}")
    if state:
        queries.append(state)
    if country:
        queries.append(country)

    for query in queries:
        if query in location_cache:
            return location_cache[query]
        try:
            location = geolocator.geocode(query)
            time.sleep(1)  # Nominatim rate limit
            if location:
                coords = [round(location.longitude, 4), round(location.latitude, 4)]
                location_cache[query] = coords
                return coords
        except Exception as e:
            print(f"Geocoding failed for '{query}': {e}")
    return None

def get_network_map_data(excel_file):
    try:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        json_path = os.path.join(script_dir, "..", "pages", "network-data.json")

        # Load partner data first
        partner_data = []
        existing_data = {}
        if os.path.exists(json_path):
            with open(json_path, 'r', encoding='utf-8') as f:
                try:
                    existing_data = json.load(f)
                    partner_data = existing_data.get("partners", [])
                except json.JSONDecodeError:
                    print("⚠️ Existing JSON is invalid. Proceeding clean.")

        workbook = openpyxl.load_workbook(excel_file, data_only=True)
        try:
            sheet = workbook[PAGE_METADATA["NETWORK_MAP"]]
        except KeyError:
            print("❌ Sheet not found.")
            print(f"Available sheets: {workbook.sheetnames}")
            return

        headers = [cell.value for cell in sheet[1]]
        cleaned_headers = [str(cell).strip() if cell else '' for cell in headers]
        expected = TABS_METADATA["NETWORK_MAP"]
        if not all(col in cleaned_headers for col in expected):
            print(f"❌ Missing required columns. Found: {cleaned_headers}")
            return

        idx = {key: cleaned_headers.index(key) for key in expected}
        impact_data = []

        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            try:
                # --- Source ---
                source_partner = row[idx['Source Partner']]
                source_state = row[idx['Source Partner State']]
                source_country = row[idx['Source partner country']]

                if not source_partner:
                    print(f"⚠️ Skipping row {row_idx} due to missing source or target partner ID.")
                    continue

                source = {
                    "partner_id": [str(source_partner).lower()] if source_partner else [],
                    "icon": "strategic"
                }
                if source_state:
                    source["stateName"] = source_state
                if source_country:
                    source["countryName"] = source_country

                # source_coords = get_coordinates(source_state, source_country, [str(source_partner).lower()] if source_partner else [], partner_data)
                # if source_coords:
                #     source["coords"] = source_coords
                # else:
                #     print(f"⚠️ Skipping row {row_idx} due to missing source coords.")
                #     continue

                # --- Target ---
                target_partner = row[idx['Target Partner']]
                target_state = row[idx['Target Partner state']]
                target_country = row[idx['Target partner country']]


                if not target_partner:
                    print(f"⚠️ Skipping row {row_idx} due to missing source or target partner ID.")
                    continue

                target = {
                    "partner_id": [str(target_partner).lower()] if target_partner else [],
                    "icon": "momentum"
                }
                if target_state:
                    target["stateName"] = target_state
                if target_country:
                    target["countryName"] = target_country

                # target_coords = get_coordinates(target_state, target_country, target_partner, partner_data)
                # if target_coords:
                #     target["coords"] = target_coords
                # else:
                #     print(f"⚠️ Skipping row {row_idx} due to missing target coords.")
                #     continue

                impact_data.append({
                    "source": source,
                    "target": target,
                    "lineType": "multi-dash",
                    "curvature": 0.3,
                    "color": "#9D6B99"
                })

            except Exception as e:
                print(f"❌ Error in row {row_idx}: {e}")
                continue

        existing_data["impactData"] = impact_data

        os.makedirs(os.path.dirname(json_path), exist_ok=True)
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(existing_data, f, indent=2, ensure_ascii=False)

                     # Dynamically import gcp_access module and upload file
        gcp_access_path = os.path.join(script_dir, '..', 'cloud-scripts', 'gcp_access.py')
        spec = importlib.util.spec_from_file_location('gcp_access', gcp_access_path)
        gcp_access = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(gcp_access)

        # folder_url = gcp_access.upload_file_to_gcs_and_get_directory(
        #     bucket_name="dev-sg-dashboard",
        #     source_file_path=json_path,
        #     destination_blob_name="sg-dashboard/landing-page.json",
        #     private_key_path=private_key_path
        # )
        #     private_key_path=private_key_path
        # )

        folder_url = gcp_access.upload_file_to_gcs_and_get_directory(
            bucket_name=os.environ.get("BUCKET_NAME"),
            source_file_path=json_path,
            destination_blob_name="sg-dashboard/network-data.json"
        )

        if folder_url:
            print(f"Successfully uploaded and got public folder URL: {folder_url}")
        else:
            print("Failed to upload file to GCS. Check logs for details.")

        print(f"✅ JSON exported successfully with {len(impact_data)} records to: {json_path}")

    except Exception as e:
        print(f"❌ Unexpected error: {str(e)}")

if __name__ == "__main__":
    excel_to_json()