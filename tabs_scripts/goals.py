import openpyxl
import json
import os
import importlib.util

from constants import PAGE_METADATA,TABS_METADATA


def goals(excel_file):
    try:
        # Get the directory of the current script
        script_dir = os.path.dirname(os.path.abspath(__file__))
        print(script_dir)
        # Define the path to the JSON file
        json_path = os.path.join(script_dir, "..", "pages", "dashboard.json")


        # Open the Excel file
        workbook = openpyxl.load_workbook(excel_file, data_only=True)
        try:
            sheet = workbook[PAGE_METADATA["GOALS"]]
        except KeyError:
            print("Error: Sheet 'Data on homepage' not found in the Excel file.")
            print(f"Available sheets: {workbook.sheetnames}")
            return


        # Get headers
        headers = [cell.value for cell in sheet[1]]
        cleaned_headers = [str(cell).strip() if cell is not None else '' for cell in headers]
        expected_columns = TABS_METADATA["GOALS"]
        if not all(col in cleaned_headers for col in expected_columns):
            print(f"Error: Excel file must contain columns: {expected_columns}")
            print(f"Found: {cleaned_headers}")
            return

        # Extract data rows
        data = []
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            try:
                row_data = {
                    'label': row[cleaned_headers.index(TABS_METADATA["GOALS"][0])] or '',
                    'value': row[cleaned_headers.index(TABS_METADATA["GOALS"][1])] or '',
                }
                if isinstance(row_data['value'], float) and row_data['value'].is_integer():
                    row_data['value'] = int(row_data['value'])
                data.append(row_data)
            except Exception as e:
                print(f"Error processing row {row_idx}: {str(e)}")
                continue


        # Read the existing JSON file
        with open(json_path, 'r', encoding='utf-8') as json_file:
            raw_content = json_file.read()
            json_file.seek(0)
            try:
                json_data = json.load(json_file)
            except json.JSONDecodeError:
                try:
                    json_data = json.loads(raw_content)
                except json.JSONDecodeError:
                    json_data = []

        # Ensure json_data is a list
        if not isinstance(json_data, list):
            json_data = [json_data] if json_data else []

        # Find and update the "dashboard-metrics" object
        found = False
        for obj in json_data:
            if isinstance(obj, dict) and obj.get('type', '').strip().lower() == 'dashboard-metrics':
                print(json_data)
                obj['indicators'] = data
                found = True
                break

        # If not found, append (this is a fallback, but your JSON should already have it)
        if not found:
            json_data.append({'type': 'dashboard-metrics', 'indicators': data})


        # Write back to the file
        with open(json_path, 'w', encoding='utf-8') as json_file:
            json.dump(json_data, json_file, indent=2, ensure_ascii=False)

        # Dynamically import gcp_access module and upload file
        gcp_access_path = os.path.join(script_dir, '..', 'cloud-scripts', 'gcp_access.py')
        spec = importlib.util.spec_from_file_location('gcp_access', gcp_access_path)
        gcp_access = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(gcp_access)

        # private_key_path = os.path.join(script_dir, "..", "private-key.json")


        # folder_url = gcp_access.upload_file_to_gcs_and_get_directory(
        #     bucket_name=os.environ.get("BUCKET_NAME"),
        #     source_file_path=json_path,
        #     destination_blob_name="sg-dashboard/landing-page.json"
        # )

        # if folder_url:
        #     print(f"Successfully uploaded and got public folder URL: {folder_url}")
        # else:
        #     print("Failed to upload file to GCS. Check logs for details.")


    except Exception as e:
        print(f"Error: {str(e)}")


