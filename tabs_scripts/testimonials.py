import openpyxl
import json
import os
import importlib.util

from constants import PAGE_METADATA,TABS_METADATA


def testimonials(excel_file):
    try:
        # Get the directory of the current script
        script_dir = os.path.dirname(os.path.abspath(__file__))
        print(script_dir)
        # Define the path to the JSON file
        json_path = os.path.join(script_dir, "..", "pages", "network-health.json")


        # Open the Excel file
        workbook = openpyxl.load_workbook(excel_file, data_only=True)
        try:
            sheet = workbook[PAGE_METADATA["TESTIMONIALS"]]
        except KeyError:
            print("Error: Sheet 'Data on homepage' not found in the Excel file.")
            print(f"Available sheets: {workbook.sheetnames}")
            return


        # Get headers
        headers = [cell.value for cell in sheet[1]]
        cleaned_headers = [str(cell).strip() if cell is not None else '' for cell in headers]
        expected_columns = TABS_METADATA["TESTIMONIALS"]
        if not all(col in cleaned_headers for col in expected_columns):
            print(f"Error: Excel file must contain columns: {expected_columns}")
            print(f"Found: {cleaned_headers}")
            return

        # Extract data rows
        data = []
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            try:
                row_data = {
                    'org': row[cleaned_headers.index(TABS_METADATA["TESTIMONIALS"][0])] or '',
                    'message': row[cleaned_headers.index(TABS_METADATA["TESTIMONIALS"][1])] or '',
                    'name': row[cleaned_headers.index(TABS_METADATA["TESTIMONIALS"][2])] or '',
                    'designation': row[cleaned_headers.index(TABS_METADATA["TESTIMONIALS"][3])] or '',
                    'image': row[cleaned_headers.index(TABS_METADATA["TESTIMONIALS"][4])] or '',
                }
                data.append(row_data)
            except Exception as e:
                print(f"Error processing row {row_idx}: {str(e)}")
                continue

        print(f"Data processing row {data}")
        if os.path.exists(json_path):
            with open(json_path, 'r', encoding='utf-8') as f:
                try:
                    json_data = json.load(f)
                except json.JSONDecodeError:
                    json_data = []
        else:
            json_data = []

        # Make sure json_data is a dict
        if not isinstance(json_data, dict):
                json_data = {}
        


         # Replace testimonials section directly
        json_data["testimonials"] = {
            "type": "testimonials",
            "width": "100%",
            "position": "left",
            "title": "Testimonials",
            "showFilters": True,  # Python boolean
            "slides": data
        }

        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(json_data, f, indent=2, ensure_ascii=False)



        # Dynamically import gcp_access module and upload file
        gcp_access_path = os.path.join(script_dir, '..', 'cloud-scripts', 'gcp_access.py')
        spec = importlib.util.spec_from_file_location('gcp_access', gcp_access_path)
        gcp_access = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(gcp_access)


        folder_url = gcp_access.upload_file_to_gcs_and_get_directory(
            bucket_name=os.environ.get("BUCKET_NAME"),
            source_file_path=json_path,
            destination_blob_name="sg-dashboard/network-health.json"
        )

        if folder_url:
            print(f"Successfully uploaded and got public folder URL: {folder_url}")
        else:
            print("Failed to upload file to GCS. Check logs for details.")


    except Exception as e:
        print(f"Error: {str(e)}")


