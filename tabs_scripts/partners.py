import openpyxl
import json
import os
import re
import requests
from constants import PAGE_METADATA, TABS_METADATA
import importlib.util


def convert_drive_link_to_direct_url(link):
    if not isinstance(link, str):
        return ''
    match = re.search(r"/d/([a-zA-Z0-9_-]+)", link)
    if not match:
        match = re.search(r"id=([a-zA-Z0-9_-]+)", link)
    if match:
        file_id = match.group(1)
        return f"https://drive.google.com/uc?export=view&id={file_id}"
    return link.strip()


def download_image(file_id, save_path):
    url = f"https://drive.google.com/uc?export=download&id={file_id}"
    try:
        response = requests.get(url)
        if response.status_code == 200:
            with open(save_path, 'wb') as f:
                f.write(response.content)
            print(f"✅ Downloaded image to {save_path}")
            return True
        else:
            print(f"❌ Failed to download image for file ID {file_id}")
            return False
    except Exception as e:
        print(f"❌ Exception downloading image: {e}")
        return False




def get_partners(excel_file):
    try:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        json_path = os.path.join(script_dir, "..", "pages", "landing-page.json")
        network_data_path  = os.path.join(script_dir, "..", "pages", "network-data.json")
        network_health_json_path = os.path.join(script_dir, "..", "pages", "network-health.json")
        images_dir = os.path.join(script_dir, "temp_downloads")
        os.makedirs(images_dir, exist_ok=True)

        workbook = openpyxl.load_workbook(excel_file, data_only=True)
        try:
            sheet = workbook[PAGE_METADATA["PARTNERS"]]
        except KeyError:
            print(f"❌ Error: Sheet '{PAGE_METADATA['PARTNERS']}' not found.")
            print("Available sheets:", workbook.sheetnames)
            return

        headers = [str(cell.value).strip() if cell.value is not None else '' for cell in sheet[1]]
        expected_columns = TABS_METADATA["PARTNERS"]

        if not all(col in headers for col in expected_columns):
            print("❌ Error: Missing required columns.")
            print("Expected:", expected_columns)
            print("Found:", headers)
            return

        data = []
        allData = []
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            try:
                raw_name = row[headers.index(expected_columns[0])]
                if not raw_name or not str(raw_name).strip():
                    continue

                raw_src = row[headers.index(expected_columns[1])] or ''
                logo_url = convert_drive_link_to_direct_url(raw_src)

                file_match = re.search(r"id=([a-zA-Z0-9_-]+)", logo_url)
                if not file_match:
                    file_match = re.search(r"/d/([a-zA-Z0-9_-]+)", raw_src)
                file_id = file_match.group(1) if file_match else ''

                name_clean = str(raw_name).strip().lower()
                name_clean = re.sub(r'[^a-z0-9_-]', '', name_clean.replace(" ", "_"))
                local_filename = f"{name_clean}.jpg"
                local_path = os.path.join(images_dir, local_filename)

                if file_id:
                    if download_image(file_id, local_path):

                         gcp_access_path = os.path.join(script_dir, '..', 'cloud-scripts', 'gcp_access.py')
                         spec = importlib.util.spec_from_file_location('gcp_access', gcp_access_path)
                         gcp_access = importlib.util.module_from_spec(spec)
                         spec.loader.exec_module(gcp_access)

                         private_key_path = os.path.join(script_dir, "..", "private-key.json")

                         folder_url = gcp_access.upload_file_to_gcs_and_get_directory(
                            bucket_name=os.environ.get("BUCKET_NAME"),
                            source_file_path=local_path,
                            destination_blob_name=f"sg-dashboard/partners/{local_filename}"
                         )

                         if folder_url:
                            os.remove(local_path)
                            final_src = f"{folder_url.rstrip('/')}/{local_filename}"
                            print(f"Successfully uploaded and got public folder URL: {folder_url}, {final_src}")
                         else:
                            print("Failed to upload file to GCS. Check logs for details.")


                row_data = {
                    'id': name_clean,
                    'src': final_src if final_src else '/assets/partners/default-partner.svg',
                    'alt': name_clean,
                    'name': str(raw_name).strip(),
                    'countryName':row[headers.index(expected_columns[2])] or '',
                    'partnerState': row[headers.index(expected_columns[3])] or '',
                    'category': row[headers.index(expected_columns[4])] or '',
                    'website': row[headers.index(expected_columns[5])] or '',
                    'coordinates': [
                        row[headers.index(expected_columns[6])] or '',
                        row[headers.index(expected_columns[7])] or ''
                    ]
                }


                allData.append(row_data)

                 # ✅ Skip if 'id' already exists
                if any(p['id'] == name_clean for p in data):
                      print(f"⚠️ Skipping duplicate partner with id: {name_clean}")
                      continue
                data.append(row_data)
                final_src= ""

            except Exception as e:
                print(f"⚠️ Error processing row {row_idx}: {e}")
                continue

        if os.path.exists(json_path):
            with open(json_path, 'r', encoding='utf-8') as f:
                try:
                    json_data = json.load(f)
                except json.JSONDecodeError:
                    json_data = []
        else:
            json_data = []

        if not isinstance(json_data, list):
            json_data = [json_data]

        found = False
        for obj in json_data:
            if isinstance(obj, dict) and obj.get('type', '').strip().lower() == 'partner-logos':
                if 'partners' not in obj or not isinstance(obj['partners'], list):
                    obj['partners'] = []
                obj['partners'] = data
                found = True
                break
            

        if not found:
            json_data.append({
                "type": "partner-logos",
                "width": "100%",
                "position": "left",
                "title": "Our Network",
                "showFilters": False,
                "partners": data,
                "styles": {
                    "section": "partner-logos-section",
                    "title": "section-title",
                    "category": "partner-category",
                    "logosContainer": "logos-container",
                    "logo": "partner-logo"
                }
            })

        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(json_data, f, indent=2, ensure_ascii=False)



            # Dynamically import gcp_access module and upload file
        gcp_access_path = os.path.join(script_dir, '..', 'cloud-scripts', 'gcp_access.py')
        spec = importlib.util.spec_from_file_location('gcp_access', gcp_access_path)
        gcp_access = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(gcp_access)

        # folder_url = gcp_access.upload_file_to_gcs_and_get_directory(
        #     bucket_name="dev-sg-dashboard",
        #     source_file_path=json_path,
        #     destination_blob_name="sg-dashboard/landing-page.json",
        #     private_key_path=private_key_path
        # )
        folder_url = gcp_access.upload_file_to_gcs_and_get_directory(
            bucket_name=os.environ.get("BUCKET_NAME"),
            source_file_path=json_path,
            destination_blob_name="sg-dashboard/landing-page.json"
        )

        if folder_url:
            print(f"Successfully uploaded and got public folder URL: {folder_url}")
        else:
            print("Failed to upload file to GCS. Check logs for details.")

        print("✅ landing-page.json updated.")
        print(f"ALL data: {allData}")


        # Load existing network data
        if os.path.exists(network_data_path):
            with open(network_data_path, 'r', encoding='utf-8') as f:
                try:
                   network_data = json.load(f)
                except json.JSONDecodeError:
                   network_data = {}
        else:
            network_data = {}

        # Ensure 'partners' exists
        if 'partners' not in network_data or not isinstance(network_data['partners'], list):
            network_data['partners'] = []

        # Append all new data
        network_data['partners'] = allData

        # ✅ Preserve existing impactData if already present
        if 'impactData' not in network_data:
            network_data['impactData'] = []

        # Save updated file
        with open(network_data_path, 'w', encoding='utf-8') as f:
            json.dump(network_data, f, indent=2, ensure_ascii=False)


                # Dynamically import gcp_access module and upload file
        gcp_access_path = os.path.join(script_dir, '..', 'cloud-scripts', 'gcp_access.py')
        spec = importlib.util.spec_from_file_location('gcp_access', gcp_access_path)
        gcp_access = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(gcp_access)

        # folder_url = gcp_access.upload_file_to_gcs_and_get_directory(
        #     bucket_name="dev-sg-dashboard",
        #     source_file_path=json_path,
        #     destination_blob_name="sg-dashboard/landing-page.json",
        #     private_key_path=private_key_path
        # )
        #     private_key_path=private_key_path
        # )

        folder_url = gcp_access.upload_file_to_gcs_and_get_directory(
            bucket_name=os.environ.get("BUCKET_NAME"),
            source_file_path=network_data_path,
            destination_blob_name="sg-dashboard/network-data.json"
        )

        if folder_url:
            print(f"Successfully uploaded and got public folder URL: {folder_url}")
        else:
            print("Failed to upload file to GCS. Check logs for details.")

        print(f"✅ Added {len(allData)} partners to network-data.json (duplicates allowed).")
        



        if os.path.exists(network_health_json_path):
            with open(network_health_json_path, 'r', encoding='utf-8') as f: 
                try:
                    network_health_data = json.load(f)
                except json.JSONDecodeError:
                    network_health_data = {}
        else:
            network_health_data = {}

         # Ensure the 'sections' list exists
        if not isinstance(network_health_data, dict):
            network_health_data = {}

        if "sections" not in network_health_data or not isinstance(network_health_data["sections"], list):
            network_health_data["sections"] = []

        sections = network_health_data["sections"]

        found = False
        for obj in sections:
            if isinstance(obj, dict) and obj.get('type', '').strip().lower() == 'partner-logos':
                if 'partners' not in obj or not isinstance(obj['partners'], list):
                    obj['partners'] = []
                obj['partners'] = data
                found = True
                break

        if not found:
            sections.append({
                "type": "partner-logos",
                "width": "100%",
                "position": "left",
                "title": "Our Network",
                "showFilters": True,
                "partners": data,
                "styles": {
                    "section": "partner-logos-section",
                    "title": "section-title",
                    "category": "partner-category",
                    "logosContainer": "logos-container",
                    "logo": "partner-logo"
                    }
          })

        # Write back to file
        with open(network_health_json_path, 'w', encoding='utf-8') as f:
            json.dump(network_health_data, f, indent=2, ensure_ascii=False)



                # Dynamically import gcp_access module and upload file
        gcp_access_path = os.path.join(script_dir, '..', 'cloud-scripts', 'gcp_access.py')
        spec = importlib.util.spec_from_file_location('gcp_access', gcp_access_path)
        gcp_access = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(gcp_access)

        # folder_url = gcp_access.upload_file_to_gcs_and_get_directory(
        #     bucket_name="dev-sg-dashboard",
        #     source_file_path=json_path,
        #     destination_blob_name="sg-dashboard/landing-page.json",
        #     private_key_path=private_key_path
        # )
        #     private_key_path=private_key_path
        # )

        folder_url = gcp_access.upload_file_to_gcs_and_get_directory(
            bucket_name=os.environ.get("BUCKET_NAME"),
            source_file_path=network_health_json_path,
            destination_blob_name="sg-dashboard/network-health.json"
        )

        if folder_url:
            print(f"Successfully uploaded and got public folder URL: {folder_url}")
        else:
            print("Failed to upload file to GCS. Check logs for details.")

        print(f"✅ Added {len(allData)} partners to network-health.json.")

    except Exception as e:
        print(f"❌ Unexpected error: {e}")

